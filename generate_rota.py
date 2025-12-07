#!/usr/bin/env python3
# Author: hitesha1981@gmail.com
## To generate shift rota for 24x7 operations with 3 shifts and 7 weekly off patterns
import argparse
import csv
import json
from datetime import datetime, timedelta
import math

from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter


# -------------------------
# Helpers
# -------------------------

def parse_date(s: str) -> datetime:
    return datetime.strptime(s, "%d-%m-%Y")


def daterange(start: datetime, end: datetime):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


# A–G weekly off patterns:
# Mon=0 ... Sun=6
OFF_PATTERNS = [
    (5, 6),  # A: Sat-Sun
    (6, 0),  # B: Sun-Mon
    (0, 1),  # C: Mon-Tue
    (1, 2),  # D: Tue-Wed
    (2, 3),  # E: Wed-Thu
    (3, 4),  # F: Thu-Fri
    (4, 5),  # G: Fri-Sat
]


# -------------------------
# Load employees
# -------------------------

def load_employees(path: str):
    employees = []
    with open(path, newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            if not row.get("emp_id") or not row.get("employee_name"):
                raise ValueError("employees.csv must contain emp_id and employee_name")
            employees.append(row)

    if not employees:
        raise ValueError("employees.csv is empty")

    # starting_shift per employee: 1,2,3; if missing/invalid, assign round-robin
    start_shifts = []
    cur = 1
    for emp in employees:
        raw = (emp.get("starting_shift") or "").strip()
        try:
            s = int(raw)
            if s not in (1, 2, 3):
                raise ValueError
        except ValueError:
            s = cur
            cur = 1 + cur % 3
        start_shifts.append(s - 1)  # store 0-based internally

    return employees, start_shifts


# -------------------------
# Precompute off pattern assignments
# -------------------------

def assign_patterns(num_employees: int):
    """Spread employees evenly across 7 patterns A–G."""
    P = len(OFF_PATTERNS)
    teams = []
    for i in range(num_employees):
        teams.append(i % P)
    return teams


def compute_off_matrix(num_employees, dates, team_of_emp):
    """off[e][d] = 1 if employee e is off on day d."""
    D = len(dates)
    off = [[0] * D for _ in range(num_employees)]
    for e in range(num_employees):
        off1, off2 = OFF_PATTERNS[team_of_emp[e]]
        for d, dt in enumerate(dates):
            wd = dt.weekday()
            if wd == off1 or wd == off2:
                off[e][d] = 1
    return off


# -------------------------
# Build & solve CP model
# -------------------------

def solve(employees, start_shifts, dates, min_per_shift=1, max_off_pct=0.30):
    E = len(employees)
    D = len(dates)
    S = 3  # 3 shifts

    model = cp_model.CpModel()

    # Assign weekly off patterns deterministically (balanced)
    team_of_emp = assign_patterns(E)
    off = compute_off_matrix(E, dates, team_of_emp)

    # x[e][d][s] = 1 if employee e works shift s on day d
    x = [[[model.NewBoolVar(f"x_e{e}_d{d}_s{s}")
           for s in range(S)] for d in range(D)] for e in range(E)]

    # ---- Base constraints ----

    # 1) Daily assignment with fixed offs
    for e in range(E):
        for d in range(D):
            if off[e][d] == 1:
                for s in range(S):
                    model.Add(x[e][d][s] == 0)
            else:
                model.Add(sum(x[e][d][s] for s in range(S)) == 1)

    # 2) 5-on / 2-off + "no shift change in 5-day work run"
    # Offs are already 2 consecutive days/week via OFF_PATTERNS.
    # We now enforce: if both days are working, shifts must be same.
    for e in range(E):
        for d in range(D - 1):
            if off[e][d] == 0 and off[e][d + 1] == 0:
                for s in range(S):
                    model.Add(x[e][d][s] == x[e][d + 1][s])

    # 3) Per-day OFF cap (~30%) – here offs are fixed; we just check
    max_off = math.ceil(max_off_pct * E)
    for d in range(D):
        model.Add(sum(off[e][d] for e in range(E)) <= max_off)

    # 4) Coverage: each shift has at least min_per_shift people
    for d in range(D):
        for s in range(S):
            model.Add(sum(x[e][d][s] for e in range(E)) >= min_per_shift)

    # ---- Fairness & 28-day rotation (soft) ----

    day_diff_vars = []
    emp_diff_vars = []
    rotation_mismatch = []

    # A) Day-wise shift balancing
    for d in range(D):
        totals = []
        for s in range(S):
            t = model.NewIntVar(0, E, f"day_tot_d{d}_s{s}")
            model.Add(t == sum(x[e][d][s] for e in range(E)))
            totals.append(t)
        for i in range(S):
            for j in range(i + 1, S):
                diff = model.NewIntVar(0, E, f"day_diff_d{d}_{i}{j}")
                model.AddAbsEquality(diff, totals[i] - totals[j])
                day_diff_vars.append(diff)

    # B) Per-employee usage of shifts (each employee sees all shifts)
    for e in range(E):
        totals = []
        for s in range(S):
            t = model.NewIntVar(0, D, f"emp_tot_e{e}_s{s}")
            model.Add(t == sum(x[e][d][s] for d in range(D)))
            totals.append(t)
        for i in range(S):
            for j in range(i + 1, S):
                diff = model.NewIntVar(0, D, f"emp_diff_e{e}_{i}{j}")
                model.AddAbsEquality(diff, totals[i] - totals[j])
                emp_diff_vars.append(diff)

    # C) 28-day rotation: prefer 1→2→3 cycles
    # For each employee e, define a "target shift" per 28-day block:
    # block_idx = d // 28; target_s = (start_shift + block_idx) % 3
    # We penalize days where actual shift != target_s.
    for e in range(E):
        start_s = start_shifts[e]
        for d in range(D):
            block = d // 28
            target_s = (start_s + block) % S
            # match = 1 if we work target shift that day, else 0 (or 0 on off days)
            match = model.NewBoolVar(f"match_e{e}_d{d}")
            # If it's a working day, match == x[e][d][target_s]
            if off[e][d] == 0:
                model.Add(match == x[e][d][target_s])
            else:
                # On off days, we don't really care; constrain match=1 so
                # these don't add penalty (or we could leave them unconstrained).
                model.Add(match == 1)
            # penalty = 1 - match
            pen = model.NewIntVar(0, 1, f"rot_pen_e{e}_d{d}")
            model.Add(pen == 1 - match)
            rotation_mismatch.append(pen)

    # D) Objective: combine all three
    # weights tuned so: day balance > per-employee balance > rotation smoothness
    model.Minimize(
        10 * sum(day_diff_vars) +
        5 * sum(emp_diff_vars) +
        1 * sum(rotation_mismatch)
    )

    # ---- Solve ----
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 120.0

    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError("No feasible rota found (check min_per_shift vs headcount).")

    # Build codes[e][d] ∈ {"1","2","3","W"}
    codes = []
    for e in range(E):
        row = []
        for d in range(D):
            if off[e][d] == 1:
                row.append("W")
            else:
                assigned_s = None
                for s in range(S):
                    if solver.Value(x[e][d][s]) == 1:
                        assigned_s = s + 1
                        break
                if assigned_s is None:
                    row.append("W")
                else:
                    row.append(str(assigned_s))
        codes.append(row)

    return codes


# -------------------------
# Excel output (.xlsx)
# -------------------------

def write_xlsx(employees, dates, codes, output_file):
    wb = Workbook()
    wb.remove(wb.active)

    fills = {
        "1": PatternFill("solid", fgColor="CCFFCC"),  # light green
        "2": PatternFill("solid", fgColor="CCE5FF"),  # light blue
        "3": PatternFill("solid", fgColor="FFCCCC"),  # light red
        "W": PatternFill("solid", fgColor="FFF2CC"),  # light yellow
    }

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    month_map = {}
    for idx, dt in enumerate(dates):
        key = (dt.year, dt.month)
        month_map.setdefault(key, []).append((dt, idx))

    for (year, month), day_list in sorted(month_map.items()):
        sheet_name = f"{year}-{month:02d}"
        ws = wb.create_sheet(sheet_name)

        # Header
        header_row = ["Name"] + [dt.day for dt, _ in day_list]
        ws.append(header_row)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center

        # Employee rows
        for r_idx, emp in enumerate(employees, start=2):
            ws.cell(r_idx, 1, emp["employee_name"])
            for c_offset, (_, global_d) in enumerate(day_list, start=2):
                code = codes[r_idx - 2][global_d]
                cell = ws.cell(r_idx, c_offset, code)
                cell.fill = fills.get(code, PatternFill())
                cell.alignment = center

        # Totals
        base_row = len(employees) + 3
        total_rows = [
            ("Shift 1 Total", "1"),
            ("Shift 2 Total", "2"),
            ("Shift 3 Total", "3"),
            ("Leave Total", "W"),
        ]
        for i, (label, code) in enumerate(total_rows):
            row = base_row + i
            c0 = ws.cell(row, 1, label)
            c0.font = header_font
            for c_offset, (_, global_d) in enumerate(day_list, start=2):
                count = sum(1 for e in range(len(employees))
                            if codes[e][global_d] == code)
                c = ws.cell(row, c_offset, count)
                c.alignment = center

        # Column widths
        ws.column_dimensions["A"].width = 24
        for col_idx in range(2, len(day_list) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 4

    wb.save(output_file)


# -------------------------
# CLI
# -------------------------

def main():
    parser = argparse.ArgumentParser(description="24x7 Shift Rota Generator")
    parser.add_argument("--start-date", required=True, help="dd-mm-yyyy")
    parser.add_argument("--stop-date", required=True, help="dd-mm-yyyy")
    parser.add_argument("--employee-details", required=True,
                        help="employees.csv (emp_id, employee_name, starting_shift, ...)")
    parser.add_argument("--config", default="config.json",
                        help="config.json with min_per_shift, max_off_pct, output_file")
    args = parser.parse_args()

    start_date = parse_date(args.start_date)
    stop_date = parse_date(args.stop_date)
    if stop_date < start_date:
        raise ValueError("stop-date must be >= start-date")

    # Config
    try:
        with open(args.config) as f:
            cfg = json.load(f)
    except FileNotFoundError:
        cfg = {}

    min_per_shift = int(cfg.get("min_per_shift", 1))
    max_off_pct = float(cfg.get("max_off_pct", 0.30))
    output_file = cfg.get("output_file", "shift_rota.xlsx")

    # Force .xlsx extension
    if not output_file.lower().endswith(".xlsx"):
        output_file = output_file.rsplit(".", 1)[0] + ".xlsx"

    employees, start_shifts = load_employees(args.employee_details)
    dates = list(daterange(start_date, stop_date))

    codes = solve(
        employees=employees,
        start_shifts=start_shifts,
        dates=dates,
        min_per_shift=min_per_shift,
        max_off_pct=max_off_pct,
    )

    write_xlsx(employees, dates, codes, output_file)
    print(f"✅ Rota generated: {output_file}")


if __name__ == "__main__":
    main()
